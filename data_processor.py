import os
import openpyxl
import pandas as pd
import numpy as np
import io
import gc
import seaborn as sns
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from PIL import Image as PILImage 
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image as OpenPyXLImage
from io import BytesIO
from PIL import Image
import matplotlib.pyplot as plt
class data_process:
    def __init__(self) -> None:
        super().__init__()
    def load_csv(self, directory):
        for root, dirs, files in os.walk(directory):
            for file in files:
                if file.endswith('.csv'):
                    yield os.path.join(root, file)  # Sử dụng yield để trả về từng tệp một  
    def checkExist_ExcelFile(self,excel_file):
        if not os.path.exists(excel_file):
            wb = Workbook()
            ws = wb.active
            ws.title = "Plots"
            ws['A1'] = 'file.csv'
            ws['B1'] = 'Timestamp'
            ws['E1'] = 'Chart'

            # Center text in headers
            header_cells = ['A1', 'B1', 'E1']
            for cell in header_cells:
                ws[cell].alignment = Alignment(horizontal='center', vertical='center')

            wb.save(excel_file)
            wb.close()
        else:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active

            # Clear data from row 2 onwards to ensure fresh data is added
            if ws.max_row > 1:
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    for cell in row:
                        cell.value = None  # Clear text data
                    ws.row_dimensions[row[0].row].height = 15

                # Remove images in the worksheet
                ws._images = []  # Clear all images
            wb.save(excel_file)
            wb.close()
    def cut_data(self,file_path):

        df_raw = pd.read_csv(file_path)
        df = df_raw[['TIME','Engine Speed','LEL Detector']]
        del df_raw
            # Convert TIME column to datetime format
        df['TIME'] = pd.to_datetime(df['TIME'], format='%H:%M:%S', errors='coerce')
            # Fill NaN values with random values
        df['Engine Speed'] = df['Engine Speed'].apply(lambda x: x if pd.notna(x) else np.random.uniform(1000, 2000))
        df['LEL Detector'] = df['LEL Detector'].apply(lambda x: x if pd.notna(x) else np.random.uniform(700, 1000))
        return df
    def verify_data(self,df):
        # Kiểm tra dữ liệu không phải là số trong cột 'Engine Speed'
        invalid_engine_speed_chars = df[~df['Engine Speed'].apply(lambda x: str(x).replace('.', '', 1).isdigit())]

            #  Kiểm tra dữ liệu không phải là số trong cột 'LEL Detector'
        invalid_lel_detector_chars = df[~df['LEL Detector'].apply(lambda x: str(x).replace('.', '', 1).isdigit())]

        if not invalid_engine_speed_chars.empty:
            print("Phát hiện dữ liệu không phải là số trong cột 'Engine Speed':")
            print(invalid_engine_speed_chars)
            raise ValueError("Phát hiện dữ liệu không phải là số trong cột 'Engine Speed'. Dừng chương trình.")
            return 0

        if not invalid_lel_detector_chars.empty:
            print("Phát hiện dữ liệu không phải là số trong cột 'LEL Detector':")
            print(invalid_lel_detector_chars)
            raise ValueError("Phát hiện dữ liệu không phải là số trong cột 'LEL Detector'. Dừng chương trình.")
            return 1
        return 2
    def create_plot(self, df, bottomlimit, toplimit, save_dir, file_name):
        # Vẽ biểu đồ
        fig, ax = plt.subplots(figsize=(5.2, 3.1))
        sns.lineplot(data=df, x='TIME', y='Engine Speed', ax=ax, color='blue', label='Engine Speed')
        sns.lineplot(data=df, x='TIME', y='LEL Detector', ax=ax, color='red', label='LEL Detector')

        # Tùy chỉnh trục x
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%H:%M'))
        ax.set_xlim(df['TIME'].min(), df['TIME'].max())
        ax.xaxis.set_major_locator(mdates.MinuteLocator(interval=1))
        plt.xticks(rotation=90)

        ax.set_ylim(bottomlimit, toplimit)
        ax.set_title('Engine Speed vs LEL Detector over TIME')
        ax.set_xlabel('TIME')
        ax.set_ylabel('Value')
        plt.legend()

        # Tạo thư mục nếu chưa tồn tại
        if not os.path.exists(save_dir):
            os.makedirs(save_dir)

        # Lưu hình ảnh vào thư mục
        img_path = os.path.join(save_dir, file_name)
        plt.savefig(img_path, format='jpeg')
        plt.close(fig)  # Đóng biểu đồ để giải phóng bộ nhớ

        # Giải phóng các biến
        del fig
        del ax
        

        gc.collect()  # Dọn dẹp bộ nhớ

        return img_path
