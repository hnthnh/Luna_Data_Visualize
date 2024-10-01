import io
import numpy as np
import seaborn as sns
import matplotlib.pyplot as plt
import pandas as pd
import os
import sys
import json
import time
import gc
import re
import matplotlib
import matplotlib.dates as mdates
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import imageio

# Đặt chế độ sử dụng cho matplotlib
matplotlib.use('Agg')

def process_csv_to_excel(directory, json_file, final_excel, output_excel_prefix='plots', batch_size=50):
    # Khai báo các thông số đầu vào từ tệp JSON
    with open(json_file, 'r') as file:
        data = json.load(file)
  
    data_columns = []
    zoom_factors = []
    bottom_limits = []
    top_limits = []
    size = [5.2, 3.1]

    for item in data:
        data_columns.append(item['column_name'])
        top_limits.append(item['upper_limit'])
        bottom_limits.append(item['lower_limit'])
        zoom_factors.append(item['expan_number'])


    number_data = len(data_columns)
    colors = ['blue', 'green', 'red', 'orange', 'purple', 'brown', 'pink', 'gray',
              'yellow', 'cyan', 'magenta', 'lime', 'teal', 'violet', 'gold', 'navy',
              'maroon', 'olive', 'coral', 'indigo', 'turquoise', 'silver', 'black'] * 3

    csv_files = sorted([f for f in os.listdir(directory) if f.endswith('.csv')])

    if not csv_files:
        print("No CSV files found in the directory. Stopping the program.")
        sys.exit()

    # Các thông số định dạng thời gian
    time_formatter = mdates.DateFormatter('%H:%M')
    time_locator = mdates.MinuteLocator(interval=5)
    excel_file_index = 1
    processed_files = 0

    # Xử lý và xuất dữ liệu từ các tệp CSV
    while processed_files < len(csv_files):
  
        
        excel_file = f'{output_excel_prefix}_{excel_file_index}.xlsx'
        
        if os.path.exists(excel_file):
            wb = load_workbook(excel_file)
            ws = wb.active
            
            if ws.max_row > 1:
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    for cell in row:
                        cell.value = None
                ws._images = []
            wb.save(excel_file)
        else:
            wb = Workbook()
            ws = wb.active

        batch = csv_files[processed_files:processed_files + batch_size]
        current_row = 2

        for csv_file in batch:
            print(f"file : {processed_files}/{len(csv_files)}")
            file_path = os.path.join(directory, csv_file)
            columns_to_read = ['TIME'] + data_columns[:number_data]
            df = pd.read_csv(file_path, usecols=columns_to_read)
            df['TIME'] = pd.to_datetime(df['TIME'], format='%H:%M:%S', errors='coerce')

            for i in range(number_data):
                column_name = data_columns[i]
                zoom_factor = zoom_factors[i]
                df[column_name] = pd.to_numeric(df[column_name], errors='coerce').fillna(0) * zoom_factor
                
                invalid_data = df[df[column_name].isna()]
                if not invalid_data.empty:
                    raise ValueError(f"Non-numeric data in '{column_name}'. Exiting program.")

                fig, ax1 = plt.subplots(figsize=size)
                sns.lineplot(data=df, x='TIME', y=column_name, ax=ax1, color=colors[i], label=column_name)
                ax1.xaxis.set_major_formatter(time_formatter)
                ax1.set_xlim(df['TIME'].min(), df['TIME'].max())
                ax1.xaxis.set_major_locator(time_locator)
                plt.xticks(rotation=90)
                plt.subplots_adjust(bottom=0.3)  # Thêm khoảng cách dưới
                ax1.set_ylim(bottom_limits[i], top_limits[i])
                ax1.set_ylabel('')
                fig.canvas.draw()
                image_array = np.array(fig.canvas.renderer.buffer_rgba())
                plt.close(fig)

                img_buffer = io.BytesIO()
                imageio.imwrite(img_buffer, image_array, format='png')
                img_buffer.seek(0)
                img_openpyxl = Image(img_buffer)
                col_letter = chr(ord('C') + i - 1)  # Tính toán ký tự cột (C = 3, D = 4, ...)
                ws.add_image(img_openpyxl, f'{col_letter}{current_row}')

            # Lấy timestamp từ tên tệp
            base_filename = csv_file.split('_')[1].replace('.csv', '')
            timestamp_str = base_filename[:8] + ' ' + base_filename[8:10] + ':' + base_filename[10:]
            formatted_timestamp = pd.to_datetime(timestamp_str, format='%Y%m%d %H:%M').strftime('%Y_%m_%d %H:%M')
            
            ws[f'A{current_row}'] = csv_file
            ws[f'B{current_row}'] = formatted_timestamp
            processed_files += 1
            current_row += 1

        wb.save(excel_file)
        wb.close()
        gc.collect()
        excel_file_index += 1

    # Ghi dữ liệu vào final_plots.xlsx
    if os.path.exists(final_excel):
        wb_final = load_workbook(final_excel)
        ws_final = wb_final.active
        
        if ws_final.max_row > 1:
            for row in ws_final.iter_rows(min_row=1, max_row=ws_final.max_row):
                for cell in row:
                    cell.value = None
                    ws_final.row_dimensions[row[0].row].height = 15  # Reset row height
            ws_final._images = []
            for col in range(1, ws_final.max_column + 1):
                ws_final.column_dimensions[get_column_letter(col)].width = 10
            headers = ['file.csv', 'Timestamp'] + [f'Chart {i + 1}' for i in range(number_data)]
            header_cells = [f'{chr(65 + i)}1' for i in range(len(headers))]
            
            for i, cell in enumerate(header_cells):
                ws_final[cell] = headers[i]
                ws_final[cell].alignment = Alignment(horizontal='center', vertical='center')
            wb_final.save(final_excel)
    else:
        wb_final = Workbook()
        ws_final = wb_final.active
        ws_final.title = "All Data"
        headers = ['file.csv', 'Timestamp'] + [f'Chart {i + 1}' for i in range(number_data)]
        ws_final.append(headers)
        header_cells = [f'{get_column_letter(i + 1)}1' for i in range(len(headers))]
        
        for cell in header_cells:
            ws_final[cell].alignment = Alignment(horizontal='center', vertical='center')

    current_row = 2

    for i in range(1, excel_file_index):
        temp_file = f'{output_excel_prefix}_{i}.xlsx'
        wb_temp = load_workbook(temp_file)
        ws_temp = wb_temp.active
        
        for row in range(2, ws_temp.max_row + 1):
            ws_final[f'A{current_row}'] = ws_temp[f'A{row}'].value
            ws_final[f'B{current_row}'] = ws_temp[f'B{row}'].value
            ws_final[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws_final[f'B{current_row}'].alignment = Alignment(horizontal='center', vertical='center')

            img_counter = 0
            
            for img in ws_temp._images:
                if isinstance(img.anchor, str):
                    img_row = int(re.findall(r'\d+', img.anchor)[0])
                elif hasattr(img.anchor, '_from'):
                    img_row = img.anchor._from.row + 1
                else:
                    raise TypeError(f"Unexpected anchor type: {type(img.anchor)}")
                
                if img_row == row:
                    if img_counter < number_data:
                        col_letter = get_column_letter(3 + img_counter)  # Bắt đầu từ cột C (3)
                        ws_final.add_image(img, f'{col_letter}{current_row}')
                    img_counter += 1
            
            if img_counter > 0:
                img_height = img.height * 0.75  # Điều chỉnh kích thước ảnh
                ws_final.row_dimensions[current_row].height = img_height + 20
            
            current_row += 1

        for col in ['A', 'B']:
            lengths = [len(str(ws_temp[f'{col}{r}'].value)) for r in range(2, ws_temp.max_row + 1) if ws_temp[f'{col}{r}'].value]
            if lengths:
                max_length = max(lengths)
                ws_final.column_dimensions[col].width = max(ws_final.column_dimensions[col].width, max_length + 2)
        
        for img in ws_temp._images:
            img_width = img.width / 7.0  # Điều chỉnh kích thước
            for col_idx in range(3, 3 + number_data):  # Các cột chứa ảnh (C, D, E,...)
                col_letter = get_column_letter(col_idx)
                ws_final.column_dimensions[col_letter].width = max(ws_final.column_dimensions[col_letter].width, img_width)

    wb_final.save(final_excel)
   
    print("Đã hoàn thành! Tất cả hình ảnh đã được lưu vào final_plots.xlsx.")
    for i in range(1, excel_file_index):
        temp_file = f'plots_{i}.xlsx'
        if os.path.exists(temp_file):
            os.remove(temp_file)
            print(f"Đã xóa {temp_file}")
    print("Đã xóa các tệp con")