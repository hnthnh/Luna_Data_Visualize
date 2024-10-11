import sys
import os
import imageio
import io
import re
import gc
import pandas as pd
from PyQt6 import uic
from PyQt6.QtWidgets import QApplication, QMainWindow, QMessageBox, QFileDialog,QTableWidgetItem
from PyQt6.QtCore import QUrl,Qt
from PyQt6.QtGui import QDesktopServices,QPixmap,QIcon
import numpy as np
import matplotlib
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import seaborn as sns
import glob
import threading
from PIL import Image as PILimg
from PIL import ImageTk as PILImageTk
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment,PatternFill,Font
from openpyxl.utils import get_column_letter
from language import translations
matplotlib.use('Agg')
lock = threading.Lock()
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        # Tải giao diện từ file .ui
        ui_file_path=self.resource_path("assets/giaodien.ui")
        uic.loadUi(ui_file_path , self)
        
        self.set_screen()
        self.clear_excel_files()
        
        icon_file_path=self.resource_path("assets\img_daihatsu_logo.ico")
        self.setWindowIcon(QIcon(icon_file_path))
        self.center()
        self.current_language = "English"  # Default language
        self.setup_comboboxLanguage()
        self.resize_columns_to_fit()
        self.btn_browsefolder.clicked.disconnect()
        self.btn_add.clicked.disconnect()
        self.btn_delete.clicked.disconnect()
        self.btn_start.clicked.disconnect()
      # Kết nối nút với hàm xử lý chỉ một lần
        self.btn_browsefolder.clicked.connect(self.on_btn_browsefolder_clicked)
        self.btn_add.clicked.connect(self.on_btn_add_clicked)
        self.btn_delete.clicked.connect(self.on_btn_delete_clicked)
        self.btn_start.clicked.connect(self.on_btn_start_clicked)
    def clear_excel_files(self):
        # Đường dẫn tới thư mục gốc
        root_directory = os.getcwd()  # Hoặc sử dụng đường dẫn cụ thể nếu cần

        # Tìm tất cả các file có dạng plots_{i}.xlsx
        excel_files = glob.glob(os.path.join(root_directory, "plots_*.xlsx"))

        # Xóa từng file tìm được
        for file in excel_files:
            try:
                os.remove(file)
            except Exception as e:
                pass
    def resource_path(self,relative_path):
        """Get absolute path to resource, works for dev and for PyInstaller."""
        try:
            base_path = sys._MEIPASS
        except Exception:
            base_path = os.path.abspath(".")

        return os.path.join(base_path, relative_path)

    def center(self):
        # Get the size of the main window
        size = self.size()  # Size of the main window

        # Get the screen where the application is running
        screen = QApplication.primaryScreen()
        screen_size = screen.size()  # Get the screen size
    
        # Calculate the x and y coordinates to center the window
        x = (screen_size.width() - size.width()) // 2
        y =0
        #y = (screen_size.height() - size.height()) // 2

        # Move the window to the calculated position
        self.move(x, y)

    def setup_comboboxLanguage(self):
        self.comboBox_language.currentIndexChanged.connect(self.change_language_from_combobox)
    def change_language_from_combobox(self):
        selected_language = self.comboBox_language.currentText()
        self.change_language(selected_language)

    def change_language(self, language):
        self.current_language = language
        self.update_translations()

    def update_translations(self):
        translations_dict = translations[self.current_language]
        self.btn_browseFolder.setText(translations_dict["btn_browseFolder"])
        self.btn_start.setText(translations_dict["btn_start"])
        self.btn_add.setText(translations_dict["btn_add"])
        self.btn_delete.setText(translations_dict["btn_delete"])

    def set_screen(self):
        screen = QApplication.primaryScreen()
        screen_size = screen.size()

        logo_path=self.resource_path("assets/img_daihatsu_logo.png")
        pixmap = QPixmap(logo_path)  # Đường dẫn tới hình ảnh của bạn
        self.label.setPixmap(pixmap)

        search_path=self.resource_path("assets/search.png")
        pixmap = QPixmap(search_path)  # Đường dẫn tới hình ảnh của bạn
        scaled_pixmap = pixmap.scaled(self.label_search.size(), 
                               aspectRatioMode=Qt.AspectRatioMode.KeepAspectRatio, 
                               transformMode=Qt.TransformationMode.SmoothTransformation)
        self.label_search.setPixmap(scaled_pixmap)

        globe_path=self.resource_path("assets\goble.png")
        pixmap = QPixmap(globe_path)  # Đường dẫn tới hình ảnh của bạn
        scaled_pixmap = pixmap.scaled(self.label_goble.size(), 
                                aspectRatioMode=Qt.AspectRatioMode.KeepAspectRatio, 
                                transformMode=Qt.TransformationMode.SmoothTransformation)
        self.label_goble.setPixmap(scaled_pixmap)
        # Tính toán kích thước 70% của màn hình
        new_width = int(screen_size.width() * 0.625)
        new_height = int(screen_size.height() * 0.9)
        # Đặt kích thước cố định 70% màn hình
        self.setFixedSize(new_width, new_height)

    def on_btn_browsefolder_clicked(self):
        # Mở hộp thoại để chọn thư mục
        folder_path = QFileDialog.getExistingDirectory(self, "Chọn thư mục")
        self.folderpath = folder_path

        if folder_path:
            # Lấy danh sách file CSV trong thư mục
            csv_files = [f for f in os.listdir(folder_path) if f.endswith('.csv')]
            if csv_files:
                # Cập nhật đường dẫn vào textbox
                self.text_browsefolder.setText(folder_path)

                # Đọc file CSV đầu tiên
                first_csv_file = os.path.join(folder_path, csv_files[0])
                df = pd.read_csv(first_csv_file)
                df.drop(columns=['TIME'], inplace=True)
                # Lấy danh sách cột và thêm vào combobox
                self.comboBox_graphitems.clear()  # Xóa các mục cũ
                self.comboBox_graphitems.addItems(df.columns.tolist())  # Thêm danh sách cột vào combobox
                QMessageBox.information(self, translations[self.current_language]["msgbox_info_title"], 
                        translations[self.current_language]["msgbox_columns_extracted"].format(len(df.columns), first_csv_file))

            else:
                # QMessageBox for CSV folder error
                QMessageBox.critical(self, translations[self.current_language]["msgbox_error_title"], 
                                    translations[self.current_language]["msgbox_no_csv_files"])

    def on_btn_add_clicked(self):
        # Lấy giá trị đã chọn từ combobox
        selected_name = self.comboBox_graphitems.currentText()

        if selected_name:
            # Lấy số lượng hàng hiện tại để làm ID
            row_position = self.table_data.rowCount()

            # Sử dụng các giá trị mặc định cho các cột khác
            upper_limit = 1200
            lower_limit = 0
            expan_number = 1.0

            # Thêm hàng vào bảng
            self.table_data.insertRow(row_position)

            # Đặt giá trị cho từng cột  
            self.table_data.setItem(row_position, 0, QTableWidgetItem(selected_name))  # Cột NAME
            self.table_data.setItem(row_position, 1, QTableWidgetItem(str(upper_limit)))  # Cột UPPER LIMIT
            self.table_data.setItem(row_position, 2, QTableWidgetItem(str(lower_limit)))  # Cột LOWER LIMIT
            self.table_data.setItem(row_position, 3, QTableWidgetItem(str(expan_number)))  # Cột EXPAN NUMBER

            # Căn giữa nội dung của các ô
            for column in range(4):  # Giả sử có 4 cột
                item = self.table_data.item(row_position, column)
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)  # Căn giữa

            # Fit columns to the overall width of the table
            self.resize_columns_to_fit()

        else:
            QMessageBox.warning(self, translations[self.current_language]["msgbox_warning_title"], 
                                translations[self.current_language]["msgbox_name_not_selected"])

    def resize_columns_to_fit(self):
        self.table_data.horizontalHeader().setStyleSheet("QHeaderView::section { background-color: #B2DFEE; color: black; }")
        table_data = self.table_data  # Get the QTableWidget instance
        total_width = table_data.width()  # Get the total width of the table
        num_columns = table_data.columnCount()  # Get the number of columns
        
        if num_columns > 0:
            # Calculate the width for each column
            column_width = total_width // num_columns
            for i in range(num_columns):
                table_data.setColumnWidth(i, column_width)  # Set the width for each column
    def on_btn_delete_clicked(self):
        # Ngắt kết nối trước khi thực hiện hành động xóa
        self.btn_delete.clicked.disconnect()

        # Thực hiện hành động xóa
        # Giả sử bạn đang sử dụng QTableWidget
        current_row = self.table_data.currentRow()
        if current_row >= 0:  # Kiểm tra xem có hàng nào được chọn không
            self.table_data.removeRow(current_row)
        else:
            # QMessageBox for row delete warning
            QMessageBox.warning(self, translations[self.current_language]["msgbox_warning_title"], 
                                translations[self.current_language]["msgbox_row_delete_warning"])

        # Kết nối lại nút Delete
        self.btn_delete.clicked.connect(self.on_btn_delete_clicked)
    def on_btn_start_clicked(self):
    # Lấy dữ liệu từ bảng
        table_data = self.get_table_data()

        # Xóa sạch dữ liệu trong QListWidget trước khi thêm mới
        self.list_process.clear()

        # Duyệt qua dữ liệu của bảng và thêm từng hàng vào QListWidget
        for row_data in table_data:
            # Chuyển đổi dữ liệu hàng thành chuỗi để hiển thị trong QListWidget
            row_text = ', '.join(row_data)
            self.list_process.addItem(row_text)
        directory = self.folderpath  # Đường dẫn thư mục chứa các file CSV

        # Kiểm tra xem directory có hợp lệ không
        if not os.path.exists(directory):
            # QMessageBox for invalid folder path
            QMessageBox.warning(self, translations[self.current_language]["msgbox_error_title"], 
                                translations[self.current_language]["msgbox_invalid_folder"])
            return

        # Hỏi người dùng chọn nơi lưu file Excel đầu ra
        final_excel, _ = QFileDialog.getSaveFileName(self, "Save Excel File", "", "Excel Files (*.xlsx)")
        
        # Kiểm tra xem người dùng có chọn file không (bấm Cancel)
        if not final_excel:
            return  # Dừng toàn bộ quá trình nếu người dùng bấm Cancel

        # Gọi hàm process_csv_to_excel và truyền các tham số
        try:
            self.process_csv_to_excel(directory, final_excel)
            # QMessageBox for success
            QMessageBox.information(self, translations[self.current_language]["msgbox_info_title"], 
                                    translations[self.current_language]["msgbox_processing_success"])
            QDesktopServices.openUrl(QUrl.fromLocalFile(final_excel))
            self.clear_excel_files()
        except Exception as e:
            # Handle exceptions with translations
            QMessageBox.critical(self, translations[self.current_language]["msgbox_error_title"], 
                                translations[self.current_language]["msgbox_general_error"].format(str(e)))
            self.clear_excel_files()
       
    def get_table_data(self):
        # Tạo danh sách để lưu dữ liệu từ table
        table_data = []

        # Duyệt qua từng hàng trong QTableWidget
        row_count = self.table_data.rowCount()
        column_count = self.table_data.columnCount()

        for row in range(row_count):
            row_data = []
            for column in range(column_count):
                # Lấy dữ liệu từ mỗi ô trong hàng
                item = self.table_data.item(row, column)
                if item is not None:
                    row_data.append(item.text())  # Lưu giá trị text của ô vào danh sách
                else:
                    row_data.append('')  # Nếu ô trống, thêm chuỗi rỗng

            # Thêm dữ liệu hàng này vào table_data
            table_data.append(row_data)

        return table_data
    def process_csv_to_excel(self, directory, final_excel):
            # Lấy dữ liệu trực tiếp từ table
            data_columns = []
            zoom_factors = []
            bottom_limits = []
            top_limits = []
            size = [5.2, 3.1]
            output_excel_prefix = 'plots'
            batch_size = 50

            # Duyệt qua các hàng trong QTableWidget
            row_count = self.table_data.rowCount()

            for row in range(row_count):
                # Kiểm tra nếu ô có giá trị
                name_item = self.table_data.item(row, 0)  # Cột "Name"
                if name_item is None or name_item.text() == '':
                    # QMessageBox for empty row errors
                    QMessageBox.critical(self, translations[self.current_language]["msgbox_error_title"], 
                                        translations[self.current_language]["msgbox_empty_name_error"].format(row))
                    return
                column_name = name_item.text()  # Lưu tên

                # Kiểm tra và chuyển đổi Upper Limit
                upper_limit_item = self.table_data.item(row, 1)  # Cột "Upper Limit"
                if upper_limit_item is None or upper_limit_item.text() == '':
                    QMessageBox.critical(self, translations[self.current_language]["msgbox_error_title"], 
                                        translations[self.current_language]["msgbox_upper_limit_empty"].format(row))
                    return
                upper_limit_str = upper_limit_item.text()
                try:
                    upper_limit = int(float(upper_limit_str))  # Chuyển từ float sang int nếu cần
                except ValueError:
                    QMessageBox.critical(self, translations[self.current_language]["msgbox_error_title"], 
                                        translations[self.current_language]["msgbox_invalid_upper_limit"].format(upper_limit_str))
                    return

                # Kiểm tra và chuyển đổi Lower Limit
                lower_limit_item = self.table_data.item(row, 2)  # Cột "Lower Limit"
                if lower_limit_item is None or lower_limit_item.text() == '':
                    QMessageBox.critical(self, translations[self.current_language]["msgbox_error_title"], 
                                        translations[self.current_language]["msgbox_lower_limit_empty"].format(row))
                    return
                lower_limit_str = lower_limit_item.text()
                try:
                    lower_limit = int(float(lower_limit_str))
                except ValueError:
                    QMessageBox.critical(self, translations[self.current_language]["msgbox_error_title"], 
                                        translations[self.current_language]["msgbox_invalid_lower_limit"].format(lower_limit_str))
                    return

                # Chuyển Expan Number thành float
                expan_number_item = self.table_data.item(row, 3)  # Cột "Expan Number"
                if expan_number_item is None or expan_number_item.text() == '':
                    QMessageBox.critical(self, translations[self.current_language]["msgbox_error_title"], 
                                        translations[self.current_language]["msgbox_expan_number_empty"].format(row))
                    return

                expan_number_str = expan_number_item.text()
                try:
                    expan_number = float(expan_number_str)  # Cột "Expan Number"
                except ValueError:
                    QMessageBox.critical(self, "Error", f"Invalid Expan Number value: {expan_number_str}")
                    return

                # Thêm vào danh sách
                data_columns.append(column_name)
                top_limits.append(upper_limit)
                bottom_limits.append(lower_limit)
                zoom_factors.append(expan_number)

            # Cập nhật số lượng dữ liệu
            number_data = len(data_columns)

            colors = ['blue', 'green', 'red','black', 'orange', 'purple', 'brown', 'black', 'pink', 'cyan', 'magenta', 
            'teal', 'violet', 'gold', 'navy', 'maroon', 'olive', 'coral', 'indigo', 'turquoise', 
            'black']

            csv_files = sorted([f for f in os.listdir(directory) if f.endswith('.csv')])

            if not csv_files:
                #print("No CSV files found in the directory. Stopping the program.")
                sys.exit()

            # Các thông số định dạng thời gian
            time_formatter = mdates.DateFormatter('%H:%M')
            time_locator = mdates.MinuteLocator(interval=5)
            excel_file_index = 1
            # Reset progress bar về 0
            self.progressBar.setValue(0)

            processed_files = 0
            total_files = len(csv_files)  # Tổng số file để xử lý
            # Xử lý và xuất dữ liệu từ các tệp CSV
            with lock:
                while processed_files < len(csv_files):
                    # Cập nhật giao diện
                    excel_file = f'{output_excel_prefix}_{excel_file_index}.xlsx'
                    
                    if os.path.exists(excel_file):
                        wb = load_workbook(excel_file)
                        ws = wb.active
                        
                        if ws.max_row > 1:
                            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                                for cell in row:
                                    cell.value = None
                            ws._images = []
                        wb.save(excel_file)
                    else:
                        wb = Workbook()
                        ws = wb.active

                    batch = csv_files[processed_files:processed_files + batch_size]
                    current_row = 2

                    for csv_file in batch:
                        #print(f"file : {processed_files}/{len(csv_files)}")
                        file_path = os.path.join(directory, csv_file)
                        columns_to_read = ['TIME'] + data_columns[:number_data]
                        df = pd.read_csv(file_path, usecols=columns_to_read)
                        df['TIME'] = pd.to_datetime(df['TIME'], format='%H:%M:%S', errors='coerce')
                        j = 0 
                        for i in range(number_data):
                            j = j + 1
                            column_name = data_columns[i]
                            zoom_factor = zoom_factors[i]
                            df[column_name] = pd.to_numeric(df[column_name], errors='coerce').fillna(0) * zoom_factor
                            
                            invalid_data = df[df[column_name].isna()]
                            if not invalid_data.empty:
                                raise ValueError(f"Non-numeric data in '{column_name}'. Exiting program.")
                            if j > len(colors):
                                j = 1
                            fig, ax1 = plt.subplots(figsize=size)
                            sns.lineplot(data=df, x='TIME', y=column_name, ax=ax1, color=colors[j], label=column_name)
                            ax1.xaxis.set_major_formatter(time_formatter)
                            ax1.set_xlim(df['TIME'].min(), df['TIME'].max())
                            ax1.xaxis.set_major_locator(time_locator)
                            plt.xticks(rotation=90)
                            plt.subplots_adjust(bottom=0.3)  # Thêm khoảng cách dưới
                            ax1.set_ylim(bottom_limits[i], top_limits[i])
                            ax1.set_ylabel('')
                            fig.canvas.draw()
                            image_array = np.array(fig.canvas.renderer.buffer_rgba())
                            plt.close(fig)

                            img_buffer = io.BytesIO()
                            imageio.imwrite(img_buffer, image_array, format='png')
                            img_buffer.seek(0)
                            img_openpyxl = Image(img_buffer)
                            col_letter = chr(ord('C') + i - 1)  # Tính toán ký tự cột (C = 3, D = 4, ...)
                            ws.add_image(img_openpyxl, f'{col_letter}{current_row}')

                        # Lấy timestamp từ tên tệp
                        base_filename = csv_file.split('_')[1].replace('.csv', '')
                        timestamp_str = base_filename[:8] + ' ' + base_filename[8:10] + ':' + base_filename[10:]
                        formatted_timestamp = pd.to_datetime(timestamp_str, format='%Y%m%d %H:%M').strftime('%Y_%m_%d %H:%M')
                        
                        ws[f'A{current_row}'] = csv_file
                        ws[f'B{current_row}'] = formatted_timestamp
                        
                        processed_files += 1
                        current_row += 1
                                        # Cập nhật thanh tiến trình trong Qt6
                        self.list_process.addItem(f"CSV Files are being processed on queue: {processed_files }/{total_files}")
                        self.list_process.scrollToBottom()
                        progress_value = int((processed_files / total_files) * 100)
                        self.progressBar.setValue(progress_value)
                        QApplication.processEvents()  # Đảm bảo giao diện người dùng được cập nhật


                    wb.save(excel_file)
                    wb.close()
                    gc.collect()
                    excel_file_index += 1
            with lock:
                # Ghi dữ liệu vào final_plots.xlsx
                if os.path.exists(final_excel):
                    wb_final = load_workbook(final_excel)
                    ws_final = wb_final.active
                    
                    if ws_final.max_row > 1:
                        for row in ws_final.iter_rows(min_row=1, max_row=ws_final.max_row):
                            for cell in row:
                                cell.value = None
                                ws_final.row_dimensions[row[0].row].height = 15  # Reset row height
                        ws_final._images = []
                        for col in range(1, ws_final.max_column + 1):
                            ws_final.column_dimensions[get_column_letter(col)].width = 10
                        headers = ['file.csv', 'Timestamp'] + [f'Chart {i + 1}' for i in range(number_data)]
                        header_cells = [f'{chr(65 + i)}1' for i in range(len(headers))]
                        
                        for i, cell in enumerate(header_cells):
                            ws_final[cell] = headers[i]
                            ws_final[cell].alignment = Alignment(horizontal='center', vertical='center')
                            ws_final[cell].font = Font(bold=True)  # Apply bold font
                            ws_final[cell].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')# Apply yellow color
                        wb_final.save(final_excel)
                else:
                    wb_final = Workbook()
                    ws_final = wb_final.active
                    ws_final.title = "All Data"
                    headers = ['file.csv', 'Timestamp'] + [f'Chart {i + 1}' for i in range(number_data)]
                    ws_final.append(headers)
                    header_cells = [f'{get_column_letter(i + 1)}1' for i in range(len(headers))]
                    
                    for cell in header_cells:
                        ws_final[cell].alignment = Alignment(horizontal='center', vertical='center')
                        ws_final[cell].font = Font(bold=True)  # Apply bold font
                        ws_final[cell].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')# Apply yellow color

            current_row = 2
            with lock:
                for i in range(1, excel_file_index):
                    temp_file = f'{output_excel_prefix}_{i}.xlsx'
                    wb_temp = load_workbook(temp_file)
                    ws_temp = wb_temp.active
                    
                    for row in range(2, ws_temp.max_row + 1):
                        ws_final[f'A{current_row}'] = ws_temp[f'A{row}'].value
                        ws_final[f'B{current_row}'] = ws_temp[f'B{row}'].value
                        ws_final[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
                        ws_final[f'B{current_row}'].alignment = Alignment(horizontal='center', vertical='center')

                        img_counter = 0
                        
                        for img in ws_temp._images:
                            if isinstance(img.anchor, str):
                                img_row = int(re.findall(r'\d+', img.anchor)[0])
                            elif hasattr(img.anchor, '_from'):
                                img_row = img.anchor._from.row + 1
                            else:
                                raise TypeError(f"Unexpected anchor type: {type(img.anchor)}")
                            
                            if img_row == row:
                                if img_counter < number_data:
                                    col_letter = get_column_letter(3 + img_counter)  # Bắt đầu từ cột C (3)
                                    ws_final.add_image(img, f'{col_letter}{current_row}')
                                img_counter += 1
                        
                        if img_counter > 0:
                            img_height = img.height * 0.75  # Điều chỉnh kích thước ảnh
                            ws_final.row_dimensions[current_row].height = img_height + 20
                        
                        current_row += 1

                    for col in ['A', 'B']:
                        lengths = [len(str(ws_temp[f'{col}{r}'].value)) for r in range(2, ws_temp.max_row + 1) if ws_temp[f'{col}{r}'].value]
                        if lengths:
                            max_length = max(lengths)
                            ws_final.column_dimensions[col].width = max(ws_final.column_dimensions[col].width, max_length + 2)
                    
                    for img in ws_temp._images:
                        img_width = img.width / 7.0  # Điều chỉnh kích thước
                        for col_idx in range(3, 3 + number_data):  # Các cột chứa ảnh (C, D, E,...)
                            col_letter = get_column_letter(col_idx)
                            ws_final.column_dimensions[col_letter].width = max(ws_final.column_dimensions[col_letter].width, img_width)

                    temp_file = f'plots_{i}.xlsx'
                    if os.path.exists(temp_file):
                        os.remove(temp_file)
            wb_final.save(final_excel)
    def change_language(self, lang):
        self.btn_add.setText(translations[lang]['btn_add'])
        self.btn_browsefolder.setText(translations[lang]['btn_browseFolder'])
        self.btn_delete.setText(translations[lang]['btn_delete'])
        self.btn_start.setText(translations[lang]['btn_start'])
        self.label_graphitems.setText(translations[lang]['label_GraphItem'])
        self.table_data.setHorizontalHeaderLabels([
        translations[lang]["tree_name"],
        translations[lang]["label_UpperLimit"],
        translations[lang]["label_LowerLimit"],
        translations[lang]["label_Expan"],
    ])


    def filter_combobox(self, text):
        # Lọc các mục trong combobox dựa trên văn bản tìm kiếm
        for index in range(self.comboBox_graphitems.count()):
            item_text = self.comboBox_graphitems.itemText(index)
            self.comboBox_graphitems.setItemData(index, item_text.lower().startswith(text.lower()), role=0)
        self.comboBox_graphitems.showPopup()    
if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setHighDpiScaleFactorRoundingPolicy(Qt.HighDpiScaleFactorRoundingPolicy.PassThrough)
    window = MainWindow()
    window.adjustSize()
    window.show()
    sys.exit(app.exec())
