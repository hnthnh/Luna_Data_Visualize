import os
import sys
import json
import csv
import shutil
import threading
import time
import gc
import re
import io
import subprocess

import numpy as np
import pandas as pd
import matplotlib
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import seaborn as sns
import imageio
import tkinter as tk


from PIL import Image as PILimg
from PIL import ImageTk as PILImageTk
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment,PatternFill,Font
from openpyxl.utils import get_column_letter
from tkinter import ttk, filedialog, messagebox
from language import translations
# Đặt chế độ sử dụng cho matplotlib
matplotlib.use('Agg')
lock = threading.Lock()
class DaihatsuApp_ver2(tk.Tk):
    def __init__(self):
        super().__init__()
       # self.destroy_temp()
        # Set window properties
        self.title("DAIHATSU DIESEL MFG.CO.,LTD | Ver 0.0.1 Beta ")
        self.iconbitmap("assets\h_logo.ico")
        window_width = 540
        window_height = 800
        self.geometry(f"{window_width}x{window_height}")
        self.resizable(False, False)
        # Get screen width and height
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        # Calculate position x, y
        position_x = int((screen_width / 2) - (window_width / 2))
        position_y = int(((screen_height+70) / 2) - ((window_height+70) / 2))
        # Set the window position
        self.geometry(f"{window_width}x{window_height}+{position_x}+{position_y}")
        self.resizable(False, False)
        # Gọi hàm setup_frame1
        # Register validation command
        self.validate_command = self.register(self.validate_number)
        self.setup_frame1()
        self.setup_frame2()
        self.setup_frame3()
        self.setup_frame4()
        self.setup_frame5()
        self.setup_frame6()
        self.load_csv()
        
        self.columns_csv = None
        # Ngôn ngữ mặc định là English
        self.current_language = "Japanese"
          
    def setup_frame1(self):
        self.frame1 = tk.Frame(self, borderwidth=2, relief="flat", padx=10, pady=10)
        self.frame1.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")

        self.image = PILimg.open("assets/img_daihatsu_logo.png")
        self.image = self.image.resize((100, 25), PILimg.LANCZOS)

        self.logo = PILImageTk.PhotoImage(self.image)
        self.logo_label = tk.Label(self.frame1, image=self.logo)
        self.logo_label.grid(row=0, column=0)

        self.browse_folder_button = ttk.Button(self.frame1, text=translations["Japanese"]["btn_browseFolder"], command=self.browse_folder)
        self.browse_folder_button.grid(row=1, column=0, padx=2, pady=2)

        # Nhập đường dẫn
        self.path_entry = tk.Entry(self.frame1,width=50)
        self.path_entry.grid(row=1, column=1, columnspan=2, padx=5, pady=5, sticky="ew")

        self.status_light = tk.Canvas(self.frame1, width=20, height=20)
        self.status_light.grid(row=1, column=3, padx=5, pady=5)
        self.update_status_light("red")  # Mặc định là màu đỏ
    def setup_frame2(self):
        self.frame2 = tk.Frame(self, borderwidth=2, relief="raised", padx=10, pady=10)
        self.frame2.grid(row=1, column=0, padx=2, pady=2, sticky="nsew")

        self.search_label = tk.Label(self.frame2, text=translations["Japanese"]["label_GraphItem"])
        self.search_label.grid(row=0, column=0)

        # Trong hàm khởi tạo hoặc nơi bạn tạo các widget:
        self.search_var = tk.StringVar()
        self.search_var.trace("w", self.update_combobox)  # Theo dõi sự thay đổi giá trị của search_var

        # Tạo Combobox
        self.combobox = ttk.Combobox(self.frame2, textvariable=self.search_var, width=20)
        self.combobox.grid(row=0, column=1, padx=0, pady=0)
       # Gán danh sách cột từ self.columns_csv vào Combobox
        self.combobox['values'] = [] # Sử dụng danh sách từ self.columns_csv

        # Tạo Button Add
        self.add_button = ttk.Button(self.frame2, text=translations["Japanese"]["btn_add"], command=self.add_column)
        self.add_button.grid(row=0, column=2)
        # Nút Delete
        self.delete_button = ttk.Button(self.frame2, text=translations["Japanese"]["btn_delete"], command=self.delete_selected)
        self.delete_button.grid(row=0, column=3)

        # Tạo Table (Treeview)
        self.tree = ttk.Treeview(self.frame2, columns=("ID", "Name"), show="headings", height=5)
        self.tree.heading("ID", text="ID")
        self.tree.heading("Name", text=translations["Japanese"]["tree_name"])
        self.tree.grid(row=1, column=0, columnspan=3, padx=5, pady=5, sticky="nsew")

        # Đặt kích thước cho cột
        self.tree.column("ID", width=100)
        self.tree.column("Name", width=200)

        # Kích thước của frame
        self.frame2.grid_rowconfigure(1, weight=1)

        # Danh sách để lưu các tên cột đã thêm
        self.column_names = []
    def setup_frame3(self):
        self.frame3 = tk.Frame(self, borderwidth=2, relief="raised", padx=10, pady=10)
        self.frame3.grid(row=2, column=0, padx=5, pady=5, sticky="nsew")
    def setup_frame4(self):
        # Khung chính của frame4
        self.frame4 = tk.Frame(self, borderwidth=2, relief="raised", padx=10, pady=10)
        self.frame4.grid(row=3, column=0, padx=10, pady=10, sticky="nsew")

        # Chia frame4 thành hai phần: frame4A và frame4B
        self.frame4A = tk.Frame(self.frame4, borderwidth=0, padx=10, pady=10)
        self.frame4A.grid(row=0, column=1, padx=10, pady=5, sticky="n")

        self.frame4B = tk.Frame(self.frame4, borderwidth=0, padx=10, pady=10)
        self.frame4B.grid(row=0, column=0, padx=10, pady=5, sticky="e")

        # Treeview cho bảng Name, Min, Max
        self.tree2 = ttk.Treeview(self.frame4B, columns=("Name", "Min", "Max"), show="headings", height=5)
        self.tree2.heading("Name", text=translations["Japanese"]["tree_name"])
        self.tree2.heading("Min", text=translations["Japanese"]["label_LowerLimit"])
        self.tree2.heading("Max", text=translations["Japanese"]["label_UpperLimit"])
        self.tree2.grid(row=1, column=0, columnspan=3, padx=5, pady=5, sticky="nsew")

        # Điều chỉnh kích thước cột
        self.tree2.column("Name", width=100)
        self.tree2.column("Min", width=70)
        self.tree2.column("Max", width=70)

        # Bind sự kiện chọn dòng trong Treeview
        self.tree2.bind("<<TreeviewSelect>>", self.on_tree2_select)

        # Tạo Label và Entry cho "Upper Limit"
        self.Upper_limit_label = tk.Label(self.frame4A, text=translations["Japanese"]["label_UpperLimit"])
        self.Upper_limit_label.grid(row=2, column=0, padx=(5, 2), pady=5, sticky="e")
        self.Upperlimit_var = tk.StringVar(value="2000")
        self.Upperlimit_entry = tk.Entry(self.frame4A, textvariable=self.Upperlimit_var, width=10, 
                                          validate='key', validatecommand=(self.validate_command, '%P'))
        self.Upperlimit_entry.grid(row=2, column=1, padx=(2, 5), pady=5, sticky="w")

        # Tạo Label và Entry cho "Lower limit"
        self.lower_limit_label = tk.Label(self.frame4A, text=translations["Japanese"]["label_LowerLimit"])
        self.lower_limit_label.grid(row=3, column=0, padx=(5, 2), pady=5, sticky="e")
        self.lower_limit_var = tk.StringVar(value="0")
        self.lower_limit_entry = tk.Entry(self.frame4A, textvariable=self.lower_limit_var, width=10, 
                                           validate='key', validatecommand=(self.validate_command, '%P'))
        self.lower_limit_entry.grid(row=3, column=1, padx=(2, 5), pady=5, sticky="w")

        # Bind sự kiện Enter cho các Entry
        self.lower_limit_entry.bind("<Return>", self.update_selected_row)
        self.Upperlimit_entry.bind("<Return>", self.update_selected_row)

        # Cấu hình cho layout giãn nở theo kích thước cửa sổ
        self.frame4.grid_columnconfigure(0, weight=1)
        self.frame4.grid_columnconfigure(1, weight=1)
        self.frame4.grid_rowconfigure(0, weight=1)
    def setup_frame5(self):
            self.frame5 = tk.Frame(self, borderwidth=2, relief="raised", padx=10, pady=10)
            self.frame5.grid(row=4, column=0, padx=5, pady=5, sticky="nsew")

            # Label và Combobox để chọn "O" hoặc "X"
            self.expan_label = tk.Label(self.frame5, text=translations["Japanese"]["label_ExpanNumber"])
            self.expan_label.grid(row=0, column=0, padx=5, pady=5, sticky="w")

            self.expan_var = tk.StringVar(value="X")  # Mặc định là "X"
            expan_option = ttk.Combobox(self.frame5, textvariable=self.expan_var, values=["O", "X"], state="readonly", width=5)
            expan_option.grid(row=0, column=1, padx=5, pady=5, sticky="e")
            expan_option.bind("<<ComboboxSelected>>", self.update_editable_state)  # Khi chọn sẽ cập nhật trạng thái khóa/mở
            
            # Treeview cho bảng Name và Expan_Number
            self.tree3 = ttk.Treeview(self.frame5, columns=("Name", "Expan_Number"), show="headings", height=5)
            self.tree3.heading("Name",  text=translations["Japanese"]["tree_name"])
            self.tree3.heading("Expan_Number", text=translations["Japanese"]["label_Expan"])
            self.tree3.grid(row=1, column=0, columnspan=2, padx=40, pady=5, sticky="nsew")

            # Điều chỉnh kích thước cột
            self.tree3.column("Name", width=150)
            self.tree3.column("Expan_Number", width=70)

            # Tạo Label và Entry cho "Expan Number"
            self.Expan_Number_label = tk.Label(self.frame5, text=translations["Japanese"]["label_Expan"])
            self.Expan_Number_label.grid(row=1, column=2, padx=(5, 2), pady=5, sticky="e")
            self.Expan_Number_var = tk.StringVar(value="1")
            
            # Thêm ràng buộc cho Entry chỉ cho phép nhập số
            self.Expan_Number_entry = tk.Entry(self.frame5, textvariable=self.Expan_Number_var, width=10, 
                                                validate='key', validatecommand=(self.validate_command, '%P'))
            self.Expan_Number_entry.grid(row=1, column=3, padx=(2, 5), pady=5, sticky="w")

            # Bind Enter event to update the selected row
            self.Expan_Number_entry.bind('<Return>', self.update_selected_row_f5)
            
            # Thêm dữ liệu từ Frame 2 vào cột Name, giá trị mặc định của Expan_Number là 1
            for item in self.tree.get_children():
                name = self.tree.item(item, "values")[1]  # Lấy giá trị từ cột Name ở Frame 2
                self.tree3.insert("", "end", values=(name, "1"))  # Giá trị mặc định của cột Expan_Number là 1

            # Gọi hàm để set trạng thái ban đầu của cột Expan_Number (khóa hoặc mở)
            self.update_editable_state()
    def setup_frame6(self):
        self.frame6 = tk.Frame(self, borderwidth=2, relief="raised", padx=10, pady=10)
        self.frame6.grid(row=5, column=0, padx=5, pady=5, sticky="nsew")

        # Biến để lưu lựa chọn ngôn ngữ
        self.language_var = tk.StringVar(value="Japanese")

        # OptionMenu cho chọn ngôn ngữ
        language_options = [ "Japanese","English"]
        language_menu = tk.OptionMenu(self.frame6, self.language_var, *language_options, command=self.update_language)
        language_menu.grid(row=0, column=0, padx=5, pady=5, sticky="w")


        # Button START ở góc phải
        self.start_button = tk.Button(self.frame6, text=translations["Japanese"]["btn_start"], font=("Arial", 15, "bold"),borderwidth=5,width=20, command=self.start_action)
        self.start_button.grid(row=0, column=2, padx=5, pady=5, sticky="ew")

        # Đặt kích thước cột
        self.frame6.grid_columnconfigure(0, weight=1)  # Cột 0 sẽ mở rộng
        self.frame6.grid_columnconfigure(1, weight=0)  # Cột 1 sẽ không mở rộng
        self.frame6.grid_columnconfigure(2, weight=0)  # Cột 2 sẽ không mở rộng
    #############################
    def get_first_csv_file(self,folder_path):
        csv_files = [f for f in os.listdir(folder_path) if f.endswith('.csv')]
        
        if csv_files:
            return os.path.join(folder_path, csv_files[0])  # Trả về đường dẫn đầy đủ của file CSV đầu tiên
        else:
            return None
    def read_csv_columns(self,file_path):
        columns = []  # Mảng để chứa tên các cột

        # Mở file CSV và đọc nội dung
        with open(file_path, mode='r', encoding='utf-8') as file:
            csv_reader = csv.reader(file)
            header = next(csv_reader)  # Lấy dòng đầu tiên làm tiêu đề (header)

            # Append các tên cột vào mảng
            columns.extend(header)

        return columns
    def browse_folder(self):
            # Mở hộp thoại để chọn thư mục
            self.folder_selected = filedialog.askdirectory()
            
            if self.folder_selected:
                # Hiển thị thư mục đã chọn
                self.path_entry.delete(0, tk.END)  # Xóa nội dung cũ
                self.path_entry.insert(0, self.folder_selected)  # Chèn đường dẫn mới vào Entry
                self.check_path()  # Kiểm tra đường dẫn ngay sau khi chọn
            else:
                messagebox.showwarning("No Folder Selected", "Please select a folder.")

             ###pick csv file and get the columns !
            csv_file_path = self.get_first_csv_file(self.folder_selected)
            self.columns_csv = self.read_csv_columns(csv_file_path)        
    def check_path(self):
        path = self.path_entry.get()
        if os.path.exists(path) and any(f.endswith('.csv') for f in os.listdir(path)):
            self.update_status_light("green")  # Đường dẫn tồn tại và có file CSV, màu xanh
        else:
            self.update_status_light("red")  # Đường dẫn không tồn tại hoặc không có file CSV, màu đỏ
    def update_status_light(self, color):
        self.status_light.delete("all")  # Xóa mọi thứ trong Canvas
        if color == "green":
            self.status_light.create_oval(2, 2, 18, 18, fill="green")  # Đèn xanh
            self.path_entry.config(state="disabled")
            

        else:
            self.status_light.create_oval(2, 2, 18, 18, fill="red")  # Đèn đỏ
    def start_action(self):
        columns_info = []

        
        # Lấy giá trị từ Treeview cho thông tin cột
        for item in self.tree2.get_children():  # tree2 chứa thông tin Name, Lower, Upper Limits
            name = self.tree2.item(item, "values")[0]  # Lấy tên cột từ Treeview
            lower_limit = self.tree2.item(item, "values")[1]  # Lấy Lower Limit
            upper_limit = self.tree2.item(item, "values")[2]  # Lấy Upper Limit
            
            # Lấy giá trị graph_limit từ Entry
            #graph_limit = self.limit_entry.get()  # Lấy giá trị Graph Limit từ Entry
            
            # Lấy Expan Number từ tree3 thay vì từ Entry
            expan_number = None
            for item3 in self.tree3.get_children():
                if self.tree3.item(item3, "values")[0] == name:
                    expan_number = self.tree3.item(item3, "values")[1]  # Lấy giá trị Expan Number từ Treeview thứ hai
            
            if expan_number is None:
                expan_number = 1  # Giá trị mặc định nếu không tìm thấy trong tree3

            # Thêm thông tin vào danh sách
            column_info = {
                'column_name': name,
                'upper_limit': int(upper_limit),
                'lower_limit': int(lower_limit),
                #'graph_limit': int(graph_limit),
                'expan_number': float(expan_number)
            }
            
            columns_info.append(column_info)

        # Lưu vào file JSON
        json_file_path = "columns_info.json"
        if os.path.exists(json_file_path):
            os.remove(json_file_path)  # Xóa file nếu nó đã tồn tại
            #(f"{json_file_path} already exists and has been deleted.")

        # Ghi lại dữ liệu vào file JSON
        with open(json_file_path, "w") as json_file:
            json.dump(columns_info, json_file, indent=4)

        #print("Data saved to columns_info.json")
        #self.output_directory = filedialog.askdirectory(title="Select Output Directory")
        self.output_excel = filedialog.asksaveasfilename(
            title="Save Excel File",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )
        # Định nghĩa các tham số
        directory = self.folder_selected  # Thay đổi đường dẫn đến thư mục chứa tệp CSV
        json_file = json_file_path  # Thay đổi đường dẫn đến tệp JSON
        output_excel_prefix = 'plots'
        final_excel = self.output_excel
        batch_size = 50
        thread = threading.Thread(target=self.process_csv_to_excel, args=(directory, json_file, final_excel, output_excel_prefix, batch_size))
        thread.start()
        #self.process_csv_to_excel(directory, json_file, final_excel, output_excel_prefix, batch_size)
    def add_column(self):
        column_name = self.search_var.get().strip()  # Lấy tên cột từ ô tìm kiếm
        if column_name and column_name in self.columns_csv:  # Kiểm tra xem có tên cột nào được nhập và có trong mảng
            if column_name not in self.column_names:
                self.column_names.append(column_name)  
                id_value = len(self.column_names) - 1  # Vị trí của phần tử mới (vị trí cuối cùng)

                # Thêm vào Treeview trên (frame 2)
                self.tree.insert("", "end", values=(id_value, column_name))
                self.search_var.set("")  # Clear search box

                # Sau khi thêm cột, gọi update_frame4_table để cập nhật bảng dưới
                self.update_frame4_table()
                self.update_frame5_table()
            else:
                messagebox.showwarning("Cảnh báo","Tên cột đã tồn tại trong danh sách.")
        else:
            messagebox.showwarning("Cảnh báo", "Tên cột không hợp lệ hoặc không có trong danh sách.")
    def delete_selected(self):
        # Lấy dòng được chọn trong Treeview
            selected_item = self.tree.selection()
            if selected_item:
                # Xóa dòng đã chọn khỏi Treeview
                self.tree.delete(selected_item)
                self.update_frame4_table()
            else:
                messagebox.showwarning("Cảnh báo", "Vui lòng chọn một dòng để xóa.")   
    def on_double_click(self):
        # Lấy mục đã chọn
        selected_item = self.tree2.selection()[0]
        values = self.tree2.item(selected_item, 'values')

        # Tạo cửa sổ để chỉnh sửa
        edit_window = tk.Toplevel(self)
        edit_window.title("Edit Limits")

        # Labels cho Lower Limit và Upper Limit
        tk.Label(edit_window, text="Lower Limit:").grid(row=0, column=0)
        tk.Label(edit_window, text="Upper Limit:").grid(row=1, column=0)

        # Entry cho Lower Limit và Upper Limit
        lower_limit_var = tk.StringVar(value=values[1])
        upper_limit_var = tk.StringVar(value=values[2])
        
        lower_limit_entry = tk.Entry(edit_window, textvariable=lower_limit_var)
        upper_limit_entry = tk.Entry(edit_window, textvariable=upper_limit_var)
        
        lower_limit_entry.grid(row=0, column=1)
        upper_limit_entry.grid(row=1, column=1)

        # Nút để xác nhận sửa đổi
        ttk.Button(edit_window, text="OK", command=lambda: self.save_changes(selected_item, lower_limit_var.get(), upper_limit_var.get(), edit_window)).grid(row=2, columnspan=2)
    def update_editable_state(self, event=None):
        # Kiểm tra lựa chọn "O" hay "X"
        if self.expan_var.get() == "X":
            # Khóa Entry khi chọn "X" và đặt lại tất cả các cột "Expan_Number" về 1
            self.Expan_Number_entry.config(state="disabled")
            for item in self.tree3.get_children():
                current_values = self.tree3.item(item, "values")
                # Đặt lại giá trị cột "Expan_Number" về 1
                self.tree3.item(item, values=(current_values[0], "1"))
        else:
            # Mở khóa Entry khi chọn "O"
            self.Expan_Number_entry.config(state="normal")
    def update_combobox(self,*args):
        search_text = self.search_var.get().lower()
        # Lọc các cột chứa từ khóa tìm kiếm
        filtered_columns = [col for col in self.columns_csv if search_text in col.lower()]
        # Cập nhật danh sách trong Combobox
        self.combobox['values'] = filtered_columns
    def update_language(self, selected_language):
        # Cập nhật ngôn ngữ hiện tại
        self.current_language = selected_language

        # Cập nhật văn bản cho các thành phần giao diện
        self.browse_folder_button.config(text=translations[self.current_language]["btn_browseFolder"])

        self.start_button.config(text=translations[self.current_language]["btn_start"])
        self.add_button.config(text=translations[self.current_language]["btn_add"])
        self.delete_button.config(text=translations[self.current_language]["btn_delete"])

        self.search_label.config(text=translations[self.current_language]["label_GraphItem"])
        self.Upper_limit_label.config(text=translations[self.current_language]["label_UpperLimit"])
        self.lower_limit_label.config(text=translations[self.current_language]["label_LowerLimit"])

        self.expan_label.config(text=translations[self.current_language]["label_ExpanNumber"])
        self.Expan_Number_label.config(text=translations[self.current_language]["label_Expan"])
        self.tree.heading("Name", text=translations[self.current_language]["tree_name"])
        self.tree2.heading("Name", text=translations[self.current_language]["tree_name"])
        self.tree2.heading("Min", text=translations[self.current_language]["label_LowerLimit"])
        self.tree2.heading("Max", text=translations[self.current_language]["label_UpperLimit"])
        self.tree3.heading("Name", text=translations[self.current_language]["tree_name"])
        self.tree3.heading("Expan_Number", text=translations[self.current_language]["label_Expan"])
    def load_csv(self):
        temp_dir='temp'
        if not os.path.exists(temp_dir):
            os.makedirs(temp_dir) 
        csv_files = [f for f in os.listdir(temp_dir) if f.endswith('.csv')]
        
        if csv_files:
            # Lấy tệp CSV đầu tiên
            csv_file_path = os.path.join(temp_dir, csv_files[0])
            try:
                with open(csv_file_path, mode='r', encoding='utf-8') as file:
                    reader = csv.reader(file)
                    columns = next(reader)  # Lấy tên cột từ dòng đầu tiên

                    # Xóa các cột hiện tại trong Treeview
                    self.tree.delete(*self.tree.get_children())
                    self.tree["columns"] = columns  # Cập nhật cột mới
                    self.tree["show"] = "headings"  # Hiện tiêu đề cột

                    # Tạo tiêu đề cột
                    for col in columns:
                        self.tree.heading(col, text=col)  # Đặt tiêu đề cột
                        self.tree.column(col, anchor="center")  # Đặt căn giữa cho cột

                    # Thêm dữ liệu vào Treeview
                    for row in reader:
                        self.tree.insert("", "end", values=row)

            except Exception as e:
                messagebox.showerror("Lỗi", f"Đã xảy ra lỗi: {e}")
        else:
            pass
    def destroy_temp(self):
        directory_path = 'temp'
        if os.path.exists(directory_path):
            try:
                shutil.rmtree(directory_path)  # Xóa thư mục và tất cả nội dung bên trong
            except Exception as e:
                pass
                #print(f"An error occurred: {e}")  # In ra thông báo lỗi nếu có
        else:
            pass
            #print("Directory does not exist.")  # Thông báo nếu thư mục không tồn tại                 
    def update_frame4_table(self):
        # Xóa các hàng hiện tại trong table frame4 trước khi update
        for item in self.tree2.get_children():
            self.tree2.delete(item)

        # Lấy dữ liệu từ bảng trên (frame 2)
        for item in self.tree.get_children():  # self.tree là Treeview của frame 2
            values = self.tree.item(item, 'values')
            column_name = values[1]  # Giả sử cột tên ở vị trí thứ 2

            # Thêm tên cột vào treeview của frame4, với Min=0 và Max=2000
            self.tree2.insert("", "end", values=(column_name, 0, 2000))

        # Bind sự kiện click vào tree2 để xử lý khi người dùng chọn hàng
        self.tree2.bind("<<TreeviewSelect>>", self.on_tree2_select)
    def update_frame5_table(self):
        # Xóa các hàng hiện tại trong table frame4 trước khi update
        for item in self.tree3.get_children():
            self.tree3.delete(item)

        # Lấy dữ liệu từ bảng trên (frame 2)
        for item in self.tree.get_children():  # self.tree là Treeview của frame 2
            values = self.tree.item(item, 'values')
            column_name = values[1]  # Giả sử cột tên ở vị trí thứ 2

            # Thêm tên cột vào treeview của frame4, với Mul = 0
            self.tree3.insert("", "end", values=(column_name, 1))

        # Bind sự kiện click vào tree2 để xử lý khi người dùng chọn hàng
        #self.tree2.bind("<<TreeviewSelect>>", self.on_tree2_select)
    def on_tree2_select(self, event):
        # Lấy hàng được chọn
        selected_item = self.tree2.selection()
        if selected_item:
            item = selected_item[0]
            values = self.tree2.item(item, "values")

            # Cập nhật Entry với giá trị Min và Max của hàng đã chọn
            self.lower_limit_var.set(values[1])  # Giá trị Lower Limit
            self.Upperlimit_var.set(values[2])   # Giá trị Upper Limit
    def update_selected_row(self, event):
        # Lấy hàng được chọn
        selected_item = self.tree2.selection()
        if selected_item:
            item = selected_item[0]

            # Lấy giá trị từ Entry
            new_min = self.lower_limit_var.get()
            new_max = self.Upperlimit_var.get()

            # Cập nhật giá trị trong Treeview
            self.tree2.item(item, values=(self.tree2.item(item, "values")[0], new_min, new_max))
    def update_selected_row_f5(self, event):

        # Lấy dòng đang được chọn
        selected_item = self.tree3.focus()
        
        if selected_item:
            # Lấy giá trị cột "Name" hiện tại
            current_values = self.tree3.item(selected_item, "values")
            
            # Cập nhật lại giá trị cột "Expan_Number" với giá trị từ Entry Expan_Number
            new_expan_value = self.Expan_Number_var.get()
            self.tree3.item(selected_item, values=(current_values[0], new_expan_value))
    def validate_number(self, value):
        """Kiểm tra xem giá trị có phải là số hợp lệ hay không (bao gồm cả float)."""
        try:
            float(value)  # Kiểm tra nếu giá trị có thể chuyển thành số float
            return True
        except ValueError:
            messagebox.showerror("Invalid input", "Please enter a valid number.")
            return False
    def process_csv_to_excel(self,directory, json_file, final_excel, output_excel_prefix='plots', batch_size=50):
        self.setup_progressWindow()  # Gọi hàm khởi tạo thanh tiến trình
        # Khai báo các thông số đầu vào từ tệp JSON
        with open(json_file, 'r') as file:
            data = json.load(file)
    
        data_columns = []
        zoom_factors = []
        bottom_limits = []
        top_limits = []
        size = [5.2, 3.1]
        for item in data:
            data_columns.append(item['column_name'])
            top_limits.append(item['upper_limit'])
            bottom_limits.append(item['lower_limit'])
            zoom_factors.append(item['expan_number'])


        number_data = len(data_columns)
        colors = ['blue', 'green', 'red','black', 'orange', 'purple', 'brown', 'black', 'pink', 'cyan', 'magenta', 
          'teal', 'violet', 'gold', 'navy', 'maroon', 'olive', 'coral', 'indigo', 'turquoise', 
          'black']

        csv_files = sorted([f for f in os.listdir(directory) if f.endswith('.csv')])

        if not csv_files:
            #print("No CSV files found in the directory. Stopping the program.")
            sys.exit()

        # Các thông số định dạng thời gian
        time_formatter = mdates.DateFormatter('%H:%M')
        time_locator = mdates.MinuteLocator(interval=5)
        excel_file_index = 1
        processed_files = 0
        # Xử lý và xuất dữ liệu từ các tệp CSV
        with lock:
            while processed_files < len(csv_files):
                  # Cập nhật giao diện
                excel_file = f'{output_excel_prefix}_{excel_file_index}.xlsx'
                
                if os.path.exists(excel_file):
                    wb = load_workbook(excel_file)
                    ws = wb.active
                    
                    if ws.max_row > 1:
                        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                            for cell in row:
                                cell.value = None
                        ws._images = []
                    wb.save(excel_file)
                else:
                    wb = Workbook()
                    ws = wb.active

                batch = csv_files[processed_files:processed_files + batch_size]
                current_row = 2

                for csv_file in batch:
                    #print(f"file : {processed_files}/{len(csv_files)}")
                    file_path = os.path.join(directory, csv_file)
                    columns_to_read = ['TIME'] + data_columns[:number_data]
                    df = pd.read_csv(file_path, usecols=columns_to_read)
                    df['TIME'] = pd.to_datetime(df['TIME'], format='%H:%M:%S', errors='coerce')
                    j = 0 
                    for i in range(number_data):
                        j = j + 1
                        column_name = data_columns[i]
                        zoom_factor = zoom_factors[i]
                        df[column_name] = pd.to_numeric(df[column_name], errors='coerce').fillna(0) * zoom_factor
                        
                        invalid_data = df[df[column_name].isna()]
                        if not invalid_data.empty:
                            raise ValueError(f"Non-numeric data in '{column_name}'. Exiting program.")
                        if j > len(colors):
                            j = 1
                        fig, ax1 = plt.subplots(figsize=size)
                        sns.lineplot(data=df, x='TIME', y=column_name, ax=ax1, color=colors[j], label=column_name)
                        ax1.xaxis.set_major_formatter(time_formatter)
                        ax1.set_xlim(df['TIME'].min(), df['TIME'].max())
                        ax1.xaxis.set_major_locator(time_locator)
                        plt.xticks(rotation=90)
                        plt.subplots_adjust(bottom=0.3)  # Thêm khoảng cách dưới
                        ax1.set_ylim(bottom_limits[i], top_limits[i])
                        ax1.set_ylabel('')
                        fig.canvas.draw()
                        image_array = np.array(fig.canvas.renderer.buffer_rgba())
                        plt.close(fig)

                        img_buffer = io.BytesIO()
                        imageio.imwrite(img_buffer, image_array, format='png')
                        img_buffer.seek(0)
                        img_openpyxl = Image(img_buffer)
                        col_letter = chr(ord('C') + i - 1)  # Tính toán ký tự cột (C = 3, D = 4, ...)
                        ws.add_image(img_openpyxl, f'{col_letter}{current_row}')

                    # Lấy timestamp từ tên tệp
                    base_filename = csv_file.split('_')[1].replace('.csv', '')
                    timestamp_str = base_filename[:8] + ' ' + base_filename[8:10] + ':' + base_filename[10:]
                    formatted_timestamp = pd.to_datetime(timestamp_str, format='%Y%m%d %H:%M').strftime('%Y_%m_%d %H:%M')
                    
                    ws[f'A{current_row}'] = csv_file
                    ws[f'B{current_row}'] = formatted_timestamp
                    
                    processed_files += 1
                    current_row += 1
                    self.progress['value'] = (processed_files / len(csv_files)) * 100  # Cập nhật thanh tiến trình
                    self.progress_window.update_idletasks()

                wb.save(excel_file)
                wb.close()
                gc.collect()
                excel_file_index += 1
        with lock:
            # Ghi dữ liệu vào final_plots.xlsx
            if os.path.exists(final_excel):
                wb_final = load_workbook(final_excel)
                ws_final = wb_final.active
                
                if ws_final.max_row > 1:
                    for row in ws_final.iter_rows(min_row=1, max_row=ws_final.max_row):
                        for cell in row:
                            cell.value = None
                            ws_final.row_dimensions[row[0].row].height = 15  # Reset row height
                    ws_final._images = []
                    for col in range(1, ws_final.max_column + 1):
                        ws_final.column_dimensions[get_column_letter(col)].width = 10
                    headers = ['file.csv', 'Timestamp'] + [f'Chart {i + 1}' for i in range(number_data)]
                    header_cells = [f'{chr(65 + i)}1' for i in range(len(headers))]
                    
                    for i, cell in enumerate(header_cells):
                        ws_final[cell] = headers[i]
                        ws_final[cell].alignment = Alignment(horizontal='center', vertical='center')
                        ws_final[cell].font = Font(bold=True)  # Apply bold font
                        ws_final[cell].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')# Apply yellow color
                    wb_final.save(final_excel)
            else:
                wb_final = Workbook()
                ws_final = wb_final.active
                ws_final.title = "All Data"
                headers = ['file.csv', 'Timestamp'] + [f'Chart {i + 1}' for i in range(number_data)]
                ws_final.append(headers)
                header_cells = [f'{get_column_letter(i + 1)}1' for i in range(len(headers))]
                
                for cell in header_cells:
                    ws_final[cell].alignment = Alignment(horizontal='center', vertical='center')
                    ws_final[cell].font = Font(bold=True)  # Apply bold font
                    ws_final[cell].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')# Apply yellow color

        current_row = 2
        with lock:
            for i in range(1, excel_file_index):
                temp_file = f'{output_excel_prefix}_{i}.xlsx'
                wb_temp = load_workbook(temp_file)
                ws_temp = wb_temp.active
                
                for row in range(2, ws_temp.max_row + 1):
                    ws_final[f'A{current_row}'] = ws_temp[f'A{row}'].value
                    ws_final[f'B{current_row}'] = ws_temp[f'B{row}'].value
                    ws_final[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_final[f'B{current_row}'].alignment = Alignment(horizontal='center', vertical='center')

                    img_counter = 0
                    
                    for img in ws_temp._images:
                        if isinstance(img.anchor, str):
                            img_row = int(re.findall(r'\d+', img.anchor)[0])
                        elif hasattr(img.anchor, '_from'):
                            img_row = img.anchor._from.row + 1
                        else:
                            raise TypeError(f"Unexpected anchor type: {type(img.anchor)}")
                        
                        if img_row == row:
                            if img_counter < number_data:
                                col_letter = get_column_letter(3 + img_counter)  # Bắt đầu từ cột C (3)
                                ws_final.add_image(img, f'{col_letter}{current_row}')
                            img_counter += 1
                    
                    if img_counter > 0:
                        img_height = img.height * 0.75  # Điều chỉnh kích thước ảnh
                        ws_final.row_dimensions[current_row].height = img_height + 20
                    
                    current_row += 1

                for col in ['A', 'B']:
                    lengths = [len(str(ws_temp[f'{col}{r}'].value)) for r in range(2, ws_temp.max_row + 1) if ws_temp[f'{col}{r}'].value]
                    if lengths:
                        max_length = max(lengths)
                        ws_final.column_dimensions[col].width = max(ws_final.column_dimensions[col].width, max_length + 2)
                
                for img in ws_temp._images:
                    img_width = img.width / 7.0  # Điều chỉnh kích thước
                    for col_idx in range(3, 3 + number_data):  # Các cột chứa ảnh (C, D, E,...)
                        col_letter = get_column_letter(col_idx)
                        ws_final.column_dimensions[col_letter].width = max(ws_final.column_dimensions[col_letter].width, img_width)

                temp_file = f'plots_{i}.xlsx'
                if os.path.exists(temp_file):
                    os.remove(temp_file)
        wb_final.save(final_excel)
    
        messagebox.showinfo("Finish", f"Excel file saved at {final_excel}.!")

        self.finish_processing()


    def setup_progressWindow(self):
            # Tạo cửa sổ thanh tiến trình
            self.progress_window = tk.Toplevel()
            self.progress_window.title("Processing...")
            self.progress_window.geometry("300x100")

            self.progress_label = tk.Label(self.progress_window, text="Processing files...")
            self.progress_label.pack(pady=10)

            self.progress = ttk.Progressbar(self.progress_window, orient="horizontal", length=250, mode="determinate")
            self.progress.pack(pady=10)
        
    def update_progress(self, value):
        self.progress['value'] = value
        self.progress_window.update_idletasks()  # Cập nhật giao diện

    def finish_processing(self):
        """Đóng cửa sổ thanh tiến trình."""
        self.progress_window.destroy()