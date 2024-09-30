import io
import numpy as np
import seaborn as sns
import matplotlib.pyplot as plt
import pandas as pd
import os
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
import sys
import matplotlib.dates as mdates
import gc
import time
import matplotlib
import imageio
import re
from openpyxl.utils import get_column_letter
import json

matplotlib.use('Agg')


class PlotExporter:
    def __init__(self, input_directory, output_excel='final_plots.xlsx', batch_size=50):
        self.input_directory = input_directory
        self.output_excel = output_excel
        self.batch_size = batch_size
        self.data_columns = []
        self.zoom_factors = []
        self.bottom_limits = []
        self.top_limits = []
        self.colors = self._generate_colors()
        self.load_column_info()
        excel_file_index = None
    def _generate_colors(self):
        return ['blue', 'green', 'red', 'orange', 'purple', 'brown', 'pink', 'gray',
                'yellow', 'cyan', 'magenta', 'lime', 'teal', 'violet', 'gold', 'navy']

    def load_column_info(self):
        with open('columns_info.json', 'r') as file:
            data = json.load(file)
        for item in data:
            self.data_columns.append(item['column_name'])
            self.top_limits.append(item['upper_limit'])
            self.bottom_limits.append(item['lower_limit'])
            self.zoom_factors.append(item['expan_number'])

    def process_files(self):
        csv_files = [f for f in os.listdir(self.input_directory) if f.endswith('.csv')]
        csv_files.sort()
        total_files = len(csv_files)
        
        if total_files == 0:
            print("No CSV files found in the directory. Stopping the program.")
            sys.exit()

        excel_file_index = 1
        processed_files = 0

        while processed_files < total_files:
            self.create_excel_file(excel_file_index)
            batch = csv_files[processed_files:processed_files + self.batch_size]
            self.create_plots(batch, processed_files)
            excel_file_index += 1

    def create_excel_file(self, index):
        excel_file = f'plots_{index}.xlsx'
        if os.path.exists(excel_file):
            self.wb = load_workbook(excel_file)
            self.ws = self.wb.active
            if self.ws.max_row > 1:
                for row in self.ws.iter_rows(min_row=2, max_row=self.ws.max_row):
                    for cell in row:
                        cell.value = None
                self.ws._images = []
            self.wb.save(excel_file)
        else:
            self.wb = Workbook()
            self.ws = self.wb.active

    def create_plots(self, batch, processed_files):
        current_row = 2
        total_files = len(batch)

        for csv_file in batch:
            file_path = os.path.join(self.input_directory, csv_file)
            columns_to_read = ['TIME'] + self.data_columns
            df = pd.read_csv(file_path, usecols=columns_to_read)
            df['TIME'] = pd.to_datetime(df['TIME'], format='%H:%M:%S', errors='coerce')

            for i, column_name in enumerate(self.data_columns):
                self.create_plot(df, column_name, i, current_row)

            base_filename = csv_file.split('_')[1].replace('.csv', '')
            timestamp_str = base_filename[:8] + ' ' + base_filename[8:10] + ':' + base_filename[10:]
            formatted_timestamp = pd.to_datetime(timestamp_str, format='%Y%m%d %H:%M').strftime('%Y_%m_%d %H:%M')
            self.ws[f'A{current_row}'] = csv_file
            self.ws[f'B{current_row}'] = formatted_timestamp
            processed_files += 1
            progress = (processed_files / total_files) * 100
            print(f"Processed {processed_files}/{total_files} file(s) ({progress:.2f}% completed).")
            current_row += 1

        self.wb.save(f'plots_{processed_files // self.batch_size}.xlsx')
        self.wb.close()
        gc.collect()

    def create_plot(self, df, column_name, index, current_row):
        zoom_factor = self.zoom_factors[index]
        df[column_name] = pd.to_numeric(df[column_name], errors='coerce').fillna(0) * zoom_factor

        if df[column_name].isna().any():
            raise ValueError(f"Non-numeric data found in '{column_name}'. Exiting program.")

        fig, ax1 = plt.subplots(figsize=(5.2, 3.1))
        sns.lineplot(data=df, x='TIME', y=column_name, ax=ax1, color=self.colors[index], label=column_name)
        ax1.set_ylim(self.bottom_limits[index], self.top_limits[index])
        plt.xticks(rotation=90)
        ax1.set_ylabel('')

        img_buffer = self.save_plot_to_image(fig)
        col_letter = chr(ord('C') + index)
        img_openpyxl = Image(img_buffer)
        self.ws.add_image(img_openpyxl, f'{col_letter}{current_row}')

    def save_plot_to_image(self, fig):
        fig.canvas.draw()
        image_array = np.array(fig.canvas.renderer.buffer_rgba())
        plt.close(fig)
        img_buffer = io.BytesIO()
        imageio.imwrite(img_buffer, image_array, format='png')
        img_buffer.seek(0)
        return img_buffer

    def finalize_excel(self):
        if os.path.exists(self.output_excel):
            self.wb_final = load_workbook(self.output_excel)
            self.ws_final = self.wb_final.active
            self.clear_previous_data()
        else:
            self.wb_final = Workbook()
            self.ws_final = self.wb_final.active
            self.ws_final.title = "All Data"

        headers = ['file.csv', 'Timestamp'] + [f'Chart {i+1}' for i in range(len(self.data_columns))]
        self.ws_final.append(headers)
        self.format_headers()

        current_row = 2
        for i in range(1, excel_file_index):
            self.add_plot_data(i, current_row)

        self.wb_final.save(self.output_excel)
        self.wb_final.close()
        print(f"Merged all plots into '{self.output_excel}'.")

    def clear_previous_data(self):
        if self.ws_final.max_row > 1:
            for row in self.ws_final.iter_rows(min_row=1, max_row=self.ws_final.max_row):
                for cell in row:
                    cell.value = None
                self.ws_final.row_dimensions[row[0].row].height = 15  # Reset row height
            self.ws_final._images = []

    def format_headers(self):
        for cell in self.ws_final[1]:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    def add_plot_data(self, index, current_row):
        temp_file = f'plots_{index}.xlsx'
        wb_temp = load_workbook(temp_file)
        ws_temp = wb_temp.active

        for row in range(2, ws_temp.max_row + 1):
            self.ws_final[f'A{current_row}'] = ws_temp[f'A{row}'].value
            self.ws_final[f'B{current_row}'] = ws_temp[f'B{row}'].value
            self.ws_final[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            self.ws_final[f'B{current_row}'].alignment = Alignment(horizontal='center', vertical='center')

            img_counter = 0
            for img in ws_temp._images:
                if isinstance(img.anchor, str):
                    img_row = int(re.findall(r'\d+', img.anchor)[0])
                elif hasattr(img.anchor, '_from'):
                    img_row = img.anchor._from.row + 1
                else:
                    raise TypeError(f"Unexpected anchor type: {type(img.anchor)}")

                if img_row == row:
                    if img_counter < len(self.data_columns):
                        col_letter = get_column_letter(3 + img_counter)  # Start from column C (3)
                        self.ws_final.add_image(img, f'{col_letter}{current_row}')
                    img_counter += 1

            if img_counter > 0:
                img_height = img.height * 0.75  # Adjust image height
                self.ws_final.row_dimensions[current_row].height = img_height + 20
            current_row += 1

        wb_temp.close()



